from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
OUT_ROOT = BASE / "곤충과제_36샘플_분석의뢰_정리표_20260609_UTF8.xlsx"
OUT_DATA = BASE / "05_데이터_표" / "곤충과제_36샘플_분석의뢰_정리표_20260609_UTF8.xlsx"


fill_title = PatternFill("solid", fgColor="1F4E78")
fill_header = PatternFill("solid", fgColor="D9EAF7")
fill_note = PatternFill("solid", fgColor="FFF2CC")
fill_yes = PatternFill("solid", fgColor="E2F0D9")
fill_hold = PatternFill("solid", fgColor="FCE4D6")
white = Font(color="FFFFFF", bold=True)
bold = Font(bold=True)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = fill_title
        cell.font = white
        cell.alignment = center
        cell.border = border


def style_body(ws) -> None:
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = wrap


def set_widths(ws, widths) -> None:
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def build_workbook() -> Workbook:
    wb = Workbook()

    ws = wb.active
    ws.title = "01_Summary"
    summary_rows = [
        ["항목", "내용"],
        [
            "연구 목적",
            "혈림프/홀바디 원물 multi-omics가 아니라, 한식연 추출공정별 동결건조 추출물의 peptide/amino acid profile 비교용 데이터 생산",
        ],
        ["시료 수", "총 36개 tube = 4종 x 9 추출조건, biological replicate 없음, 각 종-조건 조합 n=1"],
        ["시료 타입", "수용성 추출물 동결건조 분말; 혈림프/홀바디 원물 아님"],
        ["프로메타바이오 역할", "LC-MS/MS 및 amino acid 데이터 생산, 기본 QC report, raw data와 결과 테이블 제공"],
        ["중앙대 역할", "조건 비교 분석, 후보 peptide ranking, 기능성 연계 해석, 후속 validation 설계"],
        ["주요 분석", "Native peptidomics LC-MS/MS, free amino acid profiling, total amino acid profiling"],
        ["보조/선택 분석", "필요 시 수용성 extract chemistry profiling; full untargeted metabolomics는 제한적으로 해석"],
        ["보류 분석", "Lipidomics, DNA marker, RNA-seq은 혈림프/홀바디 원물 phase로 이동"],
        ["비교 축", "종 효과, 효소 효과, 초음파 효과, 초음파+효소 synergy, 열수/상온 control 비교"],
        ["필수 산출물", "Raw MS files, peptide ID/quant table, amino acid table, basic QC report, method/search parameters"],
    ]
    for row in summary_rows:
        ws.append(row)
    style_header(ws, 2)
    style_body(ws)
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).fill = fill_header
        ws.cell(r, 1).font = bold
    set_widths(ws, [24, 115])
    ws.freeze_panes = "A2"

    ws = wb.create_sheet("02_Sample_Request")
    headers = [
        "Sample_ID",
        "Tube_Label",
        "Rack",
        "Species_Code",
        "Species_KR",
        "Scientific_Name_or_State",
        "Condition_Code",
        "Extraction_Method",
        "Temperature_pH_Time",
        "Enzyme",
        "Ultrasound",
        "Expected_Product",
        "Amount_g",
        "Biological_Replicate",
        "Native_Peptidomics_Data",
        "Free_AA_Data",
        "Total_AA_Data",
        "Optional_Extract_Chemistry",
        "Vendor_Role",
        "Internal_Analysis",
        "Notes",
    ]
    ws.append(headers)

    species = [
        ("PB", "흰점박이꽃무지 유충", "Protaetia brevitarsis seulensis", "Blue"),
        ("TM", "갈색거저리/밀웜 유충", "Tenebrio molitor", "Blue"),
        ("BGJ", "백강잠", "Bombyx mori infected with Beauveria bassiana", "White"),
        ("SJ", "숙잠", "Bombyx mori, end-stage 5th instar", "White"),
    ]
    conditions = [
        ("RT", "상온", "DW 상온 추출", "room temperature", "None", "No", "수용성 baseline 추출물"),
        ("Hot", "열수", "DW 열수 추출", "90°C", "None", "No", "열수 추출물, heat-induced 변화 가능"),
        ("Pro", "Pro", "Protamex 효소 가수분해", "pH 6.0, 40°C, 18 hr", "Protamex", "No", "중간 크기 peptide"),
        ("Alc", "Alc", "Alcalase 효소 가수분해", "pH 8.0, 50°C, 18 hr", "Alcalase", "No", "다양한 peptide, endopeptidase 산물"),
        ("Fla", "Fla", "Flavourzyme 효소 가수분해", "pH 7.0, 50°C, 18 hr", "Flavourzyme", "No", "짧은 peptide + free AA"),
        ("U2H", "U2H", "초음파 단독 추출", "ultrasound 2 hr", "None", "Yes", "초음파 용출/파쇄 산물"),
        ("UPro", "UPro", "초음파 + Protamex", "ultrasound pretreatment + Protamex 18 hr", "Protamex", "Yes", "초음파-효소 복합 peptide"),
        ("UAlc", "UAlc", "초음파 + Alcalase", "ultrasound pretreatment + Alcalase 18 hr", "Alcalase", "Yes", "초음파-효소 복합 peptide"),
        ("UFla", "UFla", "초음파 + Flavourzyme", "ultrasound pretreatment + Flavourzyme 18 hr", "Flavourzyme", "Yes", "짧은 peptide + free AA 증가 예상"),
    ]
    for sp_code, sp_kr, sci, rack in species:
        for cond_code, label_suffix, method, cond_detail, enzyme, ultrasound, expected in conditions:
            sample_id = f"{sp_code}_{cond_code}"
            tube_label_kr = {"BGJ": "백강잠", "SJ": "숙잠"}.get(sp_code, sp_code)
            amount = 0.53 if sample_id == "BGJ_U2H" else 0.5
            notes = []
            if sample_id == "PB_UPro":
                notes.append("기존 표기 UPCO와 UPro 동일 조건 여부 확인")
            if sample_id == "TM_UPro":
                notes.append("기존 요약의 35개 표기와 충돌 가능: 실제 tube 존재 여부 최종 확인")
            if sample_id == "BGJ_U2H":
                notes.append("시료량 0.53 g")
            ws.append(
                [
                    sample_id,
                    f"{tube_label_kr} {label_suffix}",
                    rack,
                    sp_code,
                    sp_kr,
                    sci,
                    cond_code,
                    method,
                    cond_detail,
                    enzyme,
                    ultrasound,
                    expected,
                    amount,
                    "n=1",
                    "Yes",
                    "Yes",
                    "Yes",
                    "Optional",
                    "Data production + basic QC only",
                    "Performed internally by CAU",
                    "; ".join(notes),
                ]
            )
    style_header(ws, len(headers))
    style_body(ws)
    for row in ws.iter_rows(min_row=2):
        for c in [15, 16, 17]:
            row[c - 1].fill = fill_yes
        row[17].fill = fill_note
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    set_widths(ws, [16, 18, 10, 12, 22, 40, 14, 30, 34, 16, 14, 35, 10, 18, 22, 14, 14, 24, 30, 30, 48])

    ws = wb.create_sheet("03_Methods")
    rows = [
        ["Analysis", "Priority", "Target Samples", "Purpose", "Requested Vendor Work", "Vendor Deliverables", "Internal Work After Delivery"],
        [
            "Native peptidomics / peptide profiling LC-MS/MS",
            "1",
            "36 samples, n=1 each",
            "추출조건별 native peptide profile 비교용 데이터 생산",
            "추가 trypsin digestion 없이 endogenous/native peptide 직접 분석; DDA 우선, 가능 시 DIA 정량",
            "raw files, peptide ID table, peptide intensity table, MS/MS evidence, basic QC report",
            "조건별 peptide 비교, 후보 peptide ranking, 기능성 DB/docking/validation 설계",
        ],
        [
            "Free amino acid profiling",
            "1",
            "36 samples, n=1 each",
            "효소별 free AA release 데이터 생산",
            "HPLC/LC 기반 free AA 정량",
            "AA별 concentration, total free AA, EAA/BCAA/aromatic AA sum, QC info",
            "free AA pattern 비교, Flavourzyme/UFla 효과 해석",
        ],
        [
            "Total amino acid profiling",
            "1",
            "36 samples, n=1 each",
            "소재 기본 아미노산 조성 데이터 생산",
            "산가수분해 기반 total AA 정량",
            "total AA composition, QC info",
            "free/total ratio 계산, 영양/품질 지표 해석",
        ],
        [
            "Extract chemistry profiling",
            "Optional",
            "예산/목적에 따라 subset 또는 전체",
            "수용성 항산화 관련 저분자 또는 가공 chemistry 보조 데이터",
            "LC-MS 양/음 mode semi-targeted 또는 untargeted",
            "feature table, tentative annotation",
            "생체 metabolomics가 아닌 extract chemistry로 내부 해석",
        ],
        ["Lipidomics", "Hold", "Not requested in phase 1", "원물 phase로 이동", "Not requested", "-", "혈림프/홀바디 원물 phase에서 재설계"],
        ["DNA marker/RNA-seq", "Hold", "Not requested in phase 1", "원물 phase로 이동", "Not requested", "-", "혈림프/홀바디 원물 phase에서 재설계"],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws, 7)
    style_body(ws)
    for row in ws.iter_rows(min_row=2):
        if row[1].value == "1":
            row[1].fill = fill_yes
        elif row[1].value == "Optional":
            row[1].fill = fill_note
        else:
            row[1].fill = fill_hold
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    set_widths(ws, [34, 14, 24, 42, 55, 52, 55])

    ws = wb.create_sheet("04_QC_Design")
    rows = [
        ["항목", "권장/요청", "이유", "견적 반영 여부"],
        ["Biological replicate", "없음, 각 종-조건 조합 n=1", "교수님 지시에 따라 phase 1은 비교 screening/data production으로 진행", "견적에서 반복분 제외"],
        ["Vendor interpretation", "최소화", "해석/후보 ranking은 내부 수행", "기본 QC report까지만 요청"],
        ["Pooled QC", "가능하면 전체 시료 aliquot 소량 혼합 QC 작성", "LC-MS 안정성, drift, CV 확인", "포함 요청"],
        ["Blank", "solvent blank 및 carryover blank", "carryover/background 확인", "포함 요청"],
        ["Injection replicate", "전체 반복은 아니고 pooled QC 또는 대표 시료 일부만", "장비 재현성 확인, 비용 절감", "옵션 견적"],
        ["Randomization", "종/조건 순서를 섞어 주입", "batch/order effect 최소화", "방법에 반영 요청"],
        ["Raw data", "필수 제공", "중앙대 재분석, 후보 ranking, 후속 validation 설계", "계약/견적서에 명시"],
        ["QC report", "TIC/BPC, PCA, QC CV, missing rate", "n=1 설계에서 데이터 품질 근거로 중요", "포함 요청"],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws, 4)
    style_body(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    set_widths(ws, [24, 45, 55, 28])

    ws = wb.create_sheet("05_Vendor_Qs")
    rows = [
        ["Category", "Question", "Why it matters"],
        ["Scope", "해석이 아니라 데이터 생산 + 기본 QC report 범위로 견적 가능한가?", "조건 비교/후보 ranking은 내부 수행 예정"],
        ["Peptidomics", "추가 trypsin digestion 없이 native peptide profiling 가능한가?", "기능성 peptide가 이미 추출물 내에 존재하므로 추가 digest는 목적과 다를 수 있음"],
        ["Peptidomics", "DDA와 DIA 중 어떤 구성을 추천하며 각각 견적은?", "DDA는 ID, DIA는 조건 간 정량성에 유리"],
        ["Search", "no-enzyme/unspecific 또는 semi-specific search 가능한가?", "효소가수분해 산물은 tryptic peptide가 아님"],
        ["Search", "de novo sequencing 결과 제공 가능한가?", "Protaetia DB 등 비모델 종 DB 불완전성 보완"],
        ["Short peptides", "dipeptide/tripeptide 및 <1 kDa peptide 검출/동정 가능 범위는?", "기존 PB ACE 후보가 GF, GY, IP 등 매우 짧은 peptide 포함"],
        ["Database", "Tenebrio, Bombyx, Beauveria, Protaetia/근연종 DB 사용 가능한가?", "종별 peptide origin 추정"],
        ["Contaminants", "Bacillus/Aspergillus 효소 제제 유래 background 고려 가능한가?", "Protamex/Alcalase/Flavourzyme 유래 신호 구분"],
        ["AA", "free AA와 total AA 모두 가능한가? 가능한 AA 목록은?", "효소별 가수분해 정도와 영양/품질 평가"],
        ["QC", "pooled QC, blank, 일부 injection replicate 포함 가능한가?", "n=1 설계에서 QC 근거 필요"],
        ["Deliverables", "raw file, mzML, search parameter, FASTA, peptide table, AA table 제공 가능한가?", "중앙대 후속 재분석 및 validation 설계"],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws, 3)
    style_body(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    set_widths(ws, [18, 62, 62])

    ws = wb.create_sheet("06_References")
    rows = [
        ["Use", "Citation", "PMID", "DOI_or_Link", "Relevance"],
        ["TM 효소가수분해 AA/항산화 변화", "Tang et al., 2018, PLOS ONE", "29727456", "10.1371/journal.pone.0196218", "Tenebrio molitor에서 Alcalase/Flavourzyme 처리 후 amino acid profile과 antioxidant activity 변화"],
        ["PB ACE peptide", "Lee et al., 2023, Food Chemistry", "36037683", "10.1016/j.foodchem.2022.133897", "PB Flavourzyme hydrolysate에서 GF, GY, IP, PF, PY, SY, WI, YP, YPY 동정"],
        ["곤충 hydrolysate 비교", "Mishyna et al., 2019, Foods", "31717478", "10.3390/foods8110563", "mealworm, cricket, silkworm pupae hydrolysate 기능성 비교"],
        ["곤충 peptide discovery review", "Bioactive Peptide Discovery from Edible Insects, 2023, Molecules", "36770900", "10.3390/molecules28031233", "효소가수분해 + LC-MS/MS + 기능성 peptide discovery workflow"],
        ["곤충 bioactive peptide systematic review", "Edible Insects as a Novel Source of Bioactive Peptides, 2023, Foods", "37238844", "10.3390/foods12102026", "곤충 유래 bioactive peptide 범주와 후보 정리"],
        ["초음파+효소 원리", "Qian et al., 2023, Foods", "37959146", "10.3390/foods12214027", "ultrasound-assisted enzymatic hydrolysis mechanism and parameters"],
        ["초음파 전처리 실험근거", "Jiang et al., 2011, J Agric Food Chem", "21329351", "10.1021/jf103771x", "초음파 전처리가 효소 가수분해와 기능성 특성에 영향"],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws, 5)
    style_body(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    set_widths(ws, [30, 52, 14, 34, 65])

    ws = wb.create_sheet("07_Checklist")
    rows = [
        ["Check Item", "Status/Note"],
        ["총 샘플 수", "36개 tube 기준으로 작성"],
        ["기존 요약 35개 표기", "TM UPro 또는 기타 tube 누락/추가 여부 최종 확인 필요"],
        ["반복 설계", "biological replicate 없음, n=1 screening"],
        ["원물 정의", "혈림프/홀바디는 이번 분석 대상 아님; 현재는 수용성 추출물 동결건조 분말"],
        ["업체 의뢰 핵심", "데이터 생산 + 기본 QC report; 해석/후보 ranking은 중앙대 수행"],
        ["분석법 핵심", "일반 trypsin proteomics가 아니라 native peptidomics로 요청"],
        ["필수 deliverables", "raw file, peptide table, AA table, QC report"],
    ]
    for row in rows:
        ws.append(row)
    style_header(ws, 2)
    style_body(ws)
    set_widths(ws, [28, 95])

    for sheet in wb.worksheets:
        for row_idx in range(1, sheet.max_row + 1):
            sheet.row_dimensions[row_idx].height = 34 if row_idx > 1 else 24
    return wb


if __name__ == "__main__":
    workbook = build_workbook()
    workbook.save(OUT_ROOT)
    OUT_DATA.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUT_DATA)

    check = load_workbook(OUT_ROOT, data_only=True)
    print(OUT_ROOT)
    print(check.sheetnames)
    print("sample rows", check["02_Sample_Request"].max_row - 1)
    print(check["01_Summary"]["A1"].value, check["01_Summary"]["B2"].value[:20])
